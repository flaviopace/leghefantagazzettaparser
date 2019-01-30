import openpyxl
import sys

# La lista delle rose della tua lega
book = openpyxl.load_workbook(sys.argv[1])

sheet = book.active

cellsa = sheet['B7': 'B147']
cellsb = sheet['G7': 'G147']

giocatorirose = []

def getLegaList(array, listToAppend, enableCellinfo = False):

    for c1 in array:
        value = c1[0].value
        if value is not None and "Calciatore" not in value:
            #print value
            if not enableCellinfo:
                listToAppend.append(value)
            else:
                listToAppend.append(tuple((value, c1[0].row)))

    return listToAppend

getLegaList(cellsa, giocatorirose)
getLegaList(cellsb, giocatorirose)

print giocatorirose

# le quotazioni
book = openpyxl.load_workbook(sys.argv[2])

sheet = book.active
columnsa = sheet['C3': 'G600']

allgiocatori = []

getLegaList(columnsa, allgiocatori, True)

print allgiocatori

giocatorifree = []

for giocatore in allgiocatori:
    found = False
    for giocatorerosa in giocatorirose:
        if giocatore[0] in giocatorerosa:
            #print giocatore[0]
            found = True
    if not found:
        print "Giocatore {} non trovato ".format(giocatore[0])
        giocatorifree.append(giocatore)


print giocatorifree

wb = openpyxl.Workbook()
dest_filename = 'giocatori_liberi.xlsx'
ws = wb.active
ws.title = "Giocatori Liberi"

colname=['ID','Ruolo', 'Nome', 'Squadra', 'Quot.Att', 'Quot.Iniz', 'Diff']

colindex = 1
for val in colname:
    ws.cell(column=colindex, row=1, value="{0}".format(val))
    colindex = colindex + 1

index = 2
for giocatore in giocatorifree:

    for i in range(1,8):
        print sheet.cell(row=giocatore[1], column=i).value
        ws.cell(column=i, row=index, value="{0}".format(sheet.cell(row=giocatore[1], column=i).value))

    index = index + 1

wb.save(filename = dest_filename)

